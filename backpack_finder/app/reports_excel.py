from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from .db import get_requests, get_detections_for_request

def _autosize(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(60, max(10, max_len + 2))

def build_excel_report(limit: int = 200) -> bytes:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Requests"

    headers = ["id", "ts", "input_type", "filename", "model_yolo", "model_seg",
               "num_detections", "processing_ms", "output_image"]
    ws1.append(headers)

    requests = get_requests(limit=limit)
    for r in requests:
        ws1.append([r.get(h) for h in headers])
    _autosize(ws1)

    ws2 = wb.create_sheet("Detections")
    ws2.append([
        "request_id", "class_name", "confidence", "x1", "y1", "x2", "y2", "has_mask",
        "crop_top1_label", "crop_top1_conf", "bag_type", "bag_type_conf"
    ])

    for r in requests:
        rid = int(r["id"])
        dets = get_detections_for_request(rid)
        for d in dets:
            ws2.append([
                d["request_id"], d["class_name"], d["confidence"],
                d["x1"], d["y1"], d["x2"], d["y2"], d["has_mask"],
                d.get("crop_top1_label"), d.get("crop_top1_conf"),
                d.get("bag_type"), d.get("bag_type_conf"),
            ])
    _autosize(ws2)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
